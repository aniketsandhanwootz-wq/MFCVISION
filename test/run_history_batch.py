#!/usr/bin/env python3
from __future__ import annotations

import argparse
import concurrent.futures
import json
import mimetypes
import os
import re
import subprocess
import sys
from dataclasses import asdict, dataclass
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from html import escape
from pathlib import Path
from typing import Any
from urllib.parse import quote, urlsplit
from urllib.request import Request, urlopen
from zipfile import ZipFile
import xml.etree.ElementTree as ET

from openpyxl import Workbook, load_workbook

REPO_ROOT = Path(__file__).resolve().parents[1]
HELPER_SCRIPT = REPO_ROOT / "test" / "run_single_history_case.py"
VENV_PYTHON = REPO_ROOT / ".venv" / "bin" / "python"
CLAPPIA_FILE_DOWNLOAD_URL = (
    os.getenv("CLAPPIA_FILE_DOWNLOAD_URL") or "https://apiv2.clappia.com/file/generateFileDownloadUrl"
).strip()


NS = {"a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
CREATED_AT_RE = re.compile(r"^(\d{2}-\d{2}-\d{4})")
HYPERLINK_RE = re.compile(r'^HYPERLINK\("([^"]+)",\s*"([^"]*)"\)$', re.IGNORECASE)
RESULT_HEADERS = [
    "run_timestamp",
    "source_workbook",
    "window_start",
    "window_end",
    "batch_day",
    "day",
    "case_id",
    "submission_id",
    "owner",
    "created_at",
    "time_slot",
    "section",
    "case_label",
    "field",
    "expected_raw",
    "expected_value",
    "expected_display",
    "image_url",
    "file_key",
    "report_image_url",
    "system_status",
    "system_value",
    "system_confidence",
    "initial_match_status",
    "review_status",
    "reviewed_value",
    "review_note",
    "final_status",
    "final_value",
    "reviewed_at",
    "matched",
    "result_label",
    "error_stage",
    "reason",
    "display_kind",
    "localization_source",
    "support_read_available",
    "elapsed_seconds",
]


@dataclass(frozen=True)
class CaseRecord:
    batch_day: str
    submission_id: str
    owner: str
    created_at: str
    time_slot: str
    section: str
    case_label: str
    expected_raw: str
    expected_value: str
    expected_display: str
    image_url: str

    @property
    def case_id(self) -> str:
        return "|".join(
            [
                self.batch_day,
                self.submission_id,
                _slug(self.section),
                _slug(self.case_label),
            ]
        )


def _slug(text: str) -> str:
    cleaned = re.sub(r"[^a-z0-9]+", "_", (text or "").strip().lower())
    return cleaned.strip("_") or "unknown"


def _col_letters(col_idx: int) -> str:
    letters = []
    while col_idx > 0:
        col_idx, remainder = divmod(col_idx - 1, 26)
        letters.append(chr(65 + remainder))
    return "".join(reversed(letters))


def _parse_created_day(value: str) -> str | None:
    match = CREATED_AT_RE.match((value or "").strip())
    if not match:
        return None
    return datetime.strptime(match.group(1), "%d-%m-%Y").date().isoformat()


def _parse_hyperlink_formula(formula_text: str) -> str | None:
    if not formula_text:
        return None
    match = HYPERLINK_RE.match(formula_text.strip())
    if not match:
        return None
    return match.group(1).strip() or None


def _normalize_expected(raw_text: str) -> tuple[str | None, str | None]:
    text = (raw_text or "").strip()
    if not text:
        return None, None

    lowered = text.lower()
    if lowered in {"-infinity", "infinity", "nan"}:
        return None, None

    try:
        if re.fullmatch(r"\d+", text):
            digits = text.lstrip("0") or "0"
            if len(digits) <= 3:
                normalized = Decimal(f"0.{digits.zfill(3)}")
            else:
                normalized = Decimal(f"{digits[:-3]}.{digits[-3:]}")
        else:
            normalized = Decimal(text)
    except InvalidOperation:
        return None, None

    quantized = normalized.quantize(Decimal("0.001"))
    return format(quantized, "f"), f"{quantized:.3f}"


def _read_cell_payload(cell: ET.Element) -> dict[str, str]:
    formula = cell.find("a:f", NS)
    value = cell.find("a:v", NS)
    inline_string = cell.find("a:is", NS)
    text = ""
    if inline_string is not None:
        text = "".join(inline_string.itertext())
    elif value is not None and value.text is not None:
        text = value.text
    return {
        "formula": formula.text if formula is not None and formula.text else "",
        "value": text,
    }


def parse_cases_from_workbook(xlsx_path: Path) -> list[CaseRecord]:
    with ZipFile(xlsx_path) as zf:
        root = ET.fromstring(zf.read("xl/worksheets/sheet1.xml"))

    rows = root.findall(".//a:sheetData/a:row", NS)
    if len(rows) < 3:
        return []

    row1 = {}
    row2 = {}
    for cell in rows[0].findall("a:c", NS):
        ref = cell.attrib["r"]
        col = "".join(ch for ch in ref if ch.isalpha())
        row1[col] = _read_cell_payload(cell)["value"]
    for cell in rows[1].findall("a:c", NS):
        ref = cell.attrib["r"]
        col = "".join(ch for ch in ref if ch.isalpha())
        row2[col] = _read_cell_payload(cell)["value"]

    image_cols: list[str] = []
    for idx in range(1, max(len(row2), 123) + 1):
        col = _col_letters(idx)
        header = (row2.get(col) or "").strip()
        if header.startswith("Images of "):
            image_cols.append(col)

    cases: list[CaseRecord] = []
    for row in rows[2:]:
        payload_by_col: dict[str, dict[str, str]] = {}
        for cell in row.findall("a:c", NS):
            ref = cell.attrib["r"]
            col = "".join(ch for ch in ref if ch.isalpha())
            payload_by_col[col] = _read_cell_payload(cell)

        created_at = payload_by_col.get("C", {}).get("value", "")
        batch_day = _parse_created_day(created_at)
        if not batch_day:
            continue

        submission_id = payload_by_col.get("A", {}).get("value", "").strip()
        owner = payload_by_col.get("B", {}).get("value", "").strip()
        time_slot = payload_by_col.get("G", {}).get("value", "").strip()

        for image_col in image_cols:
            value_col = _col_letters(_letters_to_num(image_col) - 1)
            value_header = (row2.get(value_col) or "").strip()
            section = (row1.get(image_col) or row1.get(value_col) or "").strip()
            expected_raw = payload_by_col.get(value_col, {}).get("value", "").strip()
            expected_value, expected_display = _normalize_expected(expected_raw)
            image_formula = payload_by_col.get(image_col, {}).get("formula", "").strip()
            image_url = _parse_hyperlink_formula(image_formula)

            if not value_header or not expected_value or not image_url:
                continue

            cases.append(
                CaseRecord(
                    batch_day=batch_day,
                    submission_id=submission_id,
                    owner=owner,
                    created_at=created_at.strip(),
                    time_slot=time_slot,
                    section=section or "Unknown Section",
                    case_label=value_header,
                    expected_raw=expected_raw,
                    expected_value=expected_value,
                    expected_display=expected_display,
                    image_url=image_url,
                )
            )

    return cases


def _letters_to_num(letters: str) -> int:
    value = 0
    for char in letters:
        value = value * 26 + (ord(char.upper()) - 64)
    return value


def select_last_two_week_days(cases: list[CaseRecord]) -> list[str]:
    all_days = sorted({case.batch_day for case in cases})
    if not all_days:
        return []
    max_day = date.fromisoformat(all_days[-1])
    window_start = max_day - timedelta(days=13)
    return [day for day in all_days if window_start <= date.fromisoformat(day) <= max_day]


def _matched(expected_value: str | None, system_value: str | None, status: str) -> bool:
    if status != "ok" or not expected_value or not system_value:
        return False
    try:
        expected_decimal = Decimal(expected_value).quantize(Decimal("0.001"))
        system_decimal = Decimal(system_value).quantize(Decimal("0.001"))
    except InvalidOperation:
        return False
    return expected_decimal == system_decimal


def _match_status(status: str, matched: bool) -> str:
    if matched:
        return "correct"
    if status in {"fetch_failed", "execution_failed"}:
        return "blocked"
    if status == "needs_review":
        return "needs_review"
    return "mismatch"


def _is_last_decimal_only_variation(expected_value: str | None, system_value: str | None, status: str) -> bool:
    if status != "ok" or not expected_value or not system_value:
        return False
    try:
        expected_decimal = Decimal(expected_value)
        system_decimal = Decimal(system_value)
    except InvalidOperation:
        return False
    return (
        expected_decimal.quantize(Decimal("0.01")) == system_decimal.quantize(Decimal("0.01"))
        and expected_decimal.quantize(Decimal("0.001")) != system_decimal.quantize(Decimal("0.001"))
    )


def _result_label(status: str, matched: bool) -> str:
    if matched:
        return "Correct"
    if status in {"fetch_failed", "execution_failed"}:
        return "Blocked"
    if status == "needs_review":
        return "Needs review"
    if status != "ok":
        return "Failed"
    return "Mismatch"


def _run_single_case(
    case: CaseRecord,
    *,
    run_timestamp: str,
    source_workbook: str,
    window_start: str,
    window_end: str,
) -> dict[str, Any]:
    row = {
        "run_timestamp": run_timestamp,
        "source_workbook": source_workbook,
        "window_start": window_start,
        "window_end": window_end,
        "batch_day": case.batch_day,
        "day": case.batch_day,
        "case_id": case.case_id,
        "submission_id": case.submission_id,
        "owner": case.owner,
        "created_at": case.created_at,
        "time_slot": case.time_slot,
        "section": case.section,
        "case_label": case.case_label,
        "field": case.case_label,
        "expected_raw": case.expected_raw,
        "expected_value": case.expected_value,
        "expected_display": case.expected_display,
        "image_url": case.image_url,
        "file_key": "",
        "report_image_url": "",
        "system_status": "",
        "system_value": "",
        "system_confidence": "",
        "initial_match_status": "",
        "review_status": "",
        "reviewed_value": "",
        "review_note": "",
        "final_status": "",
        "final_value": "",
        "reviewed_at": "",
        "matched": "no",
        "result_label": "",
        "error_stage": "",
        "reason": "",
        "display_kind": "",
        "localization_source": "",
        "support_read_available": "no",
        "elapsed_seconds": "",
    }
    try:
        analysis = _run_analysis_helper(
            image_url=case.image_url,
            trace_id=f"history-{_slug(case.case_id)}",
            target_key=_slug(case.case_label),
        )
        final = analysis["final"]
        localization = analysis.get("localization") or {}
        support_read = analysis.get("support_read") or {}
        history_source = analysis.get("_history_source") or {}

        system_value = final.get("value_text")
        system_status = final.get("status") or "needs_review"
        is_match = _matched(case.expected_value, system_value, system_status)
        match_status = _match_status(system_status, is_match)
        row.update(
            {
                "file_key": history_source.get("file_key"),
                "report_image_url": history_source.get("fetch_image_url") or case.image_url,
                "system_status": system_status,
                "system_value": system_value,
                "system_confidence": final.get("confidence"),
                "initial_match_status": match_status,
                "final_status": match_status,
                "final_value": system_value,
                "matched": "yes" if is_match else "no",
                "result_label": _result_label(system_status, is_match),
                "reason": final.get("reason"),
                "display_kind": localization.get("display_kind"),
                "localization_source": localization.get("source"),
                "support_read_available": "yes" if support_read.get("available") else "no",
                "elapsed_seconds": analysis.get("elapsed_seconds"),
            }
        )
    except Exception as exc:
        message = str(exc)
        status = "fetch_failed" if "remote url is not an image" in message.lower() or "could not fetch image url" in message.lower() else "execution_failed"
        row.update(
            {
                "system_status": status,
                "initial_match_status": _match_status(status, False),
                "final_status": _match_status(status, False),
                "result_label": _result_label(status, False),
                "error_stage": "fetch" if status == "fetch_failed" else "execution",
                "reason": message,
            }
        )
    return row


def run_batch(
    cases: list[CaseRecord],
    *,
    source_workbook: str,
    window_start: str,
    window_end: str,
    workers: int,
) -> list[dict[str, Any]]:
    run_timestamp = datetime.now().strftime("%Y-%m-%d %I:%M:%S %p")
    max_workers = max(1, min(workers, len(cases)))
    if max_workers == 1:
        return [
            _run_single_case(
                case,
                run_timestamp=run_timestamp,
                source_workbook=source_workbook,
                window_start=window_start,
                window_end=window_end,
            )
            for case in cases
        ]

    results_by_case_id: dict[str, dict[str, Any]] = {}
    with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
        future_map = {
            executor.submit(
                _run_single_case,
                case,
                run_timestamp=run_timestamp,
                source_workbook=source_workbook,
                window_start=window_start,
                window_end=window_end,
            ): case.case_id
            for case in cases
        }
        for future in concurrent.futures.as_completed(future_map):
            case_id = future_map[future]
            results_by_case_id[case_id] = future.result()

    return [results_by_case_id[case.case_id] for case in cases]


def _run_analysis_helper(*, image_url: str, trace_id: str, target_key: str) -> dict[str, Any]:
    completed = subprocess.run(
        [
            str(VENV_PYTHON),
            str(HELPER_SCRIPT),
            "--image-url",
            image_url,
            "--trace-id",
            trace_id,
            "--target-key",
            target_key,
        ],
        cwd=str(REPO_ROOT),
        capture_output=True,
        text=True,
        timeout=600,
        check=False,
    )
    if completed.returncode != 0:
        raise RuntimeError(
            "History case helper failed.\n"
            f"stdout:\n{completed.stdout}\n"
            f"stderr:\n{completed.stderr}"
        )
    return json.loads(completed.stdout)


def load_existing_case_rows(workbook_path: Path) -> list[dict[str, Any]]:
    if not workbook_path.exists():
        return []
    wb = load_workbook(workbook_path)
    if "cases" not in wb.sheetnames:
        return []
    ws = wb["cases"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(cell) if cell is not None else "" for cell in rows[0]]
    existing: list[dict[str, Any]] = []
    for row in rows[1:]:
        if not any(cell not in (None, "") for cell in row):
            continue
        existing.append({header: row[idx] for idx, header in enumerate(headers)})
    return existing


def _normalize_existing_case_row(row: dict[str, Any]) -> dict[str, Any]:
    normalized = {header: row.get(header, "") for header in RESULT_HEADERS}
    normalized["day"] = normalized.get("day") or normalized.get("batch_day") or ""
    normalized["field"] = normalized.get("field") or normalized.get("case_label") or ""
    if not normalized["initial_match_status"]:
        normalized["initial_match_status"] = _match_status(
            str(normalized.get("system_status") or ""),
            str(normalized.get("matched") or "").lower() == "yes",
        )
    if not normalized["final_status"]:
        normalized["final_status"] = normalized["initial_match_status"]
    if not normalized["final_value"]:
        normalized["final_value"] = normalized.get("system_value") or ""
    for key in ("review_status", "reviewed_value", "review_note", "reviewed_at"):
        normalized[key] = normalized.get(key) or ""
    return normalized


def _merge_review_fields(
    new_row: dict[str, Any],
    existing_row: dict[str, Any] | None,
) -> dict[str, Any]:
    merged = dict(new_row)
    merged["day"] = merged.get("day") or merged.get("batch_day") or ""
    merged["field"] = merged.get("field") or merged.get("case_label") or ""
    if existing_row is None:
        merged.setdefault("review_status", "")
        merged.setdefault("reviewed_value", "")
        merged.setdefault("review_note", "")
        merged.setdefault("reviewed_at", "")
        merged.setdefault("final_status", merged.get("initial_match_status") or "")
        merged.setdefault("final_value", merged.get("system_value") or "")
        return merged

    existing = _normalize_existing_case_row(existing_row)
    for key in ("review_status", "reviewed_value", "review_note", "reviewed_at"):
        merged[key] = existing.get(key) or ""
    merged["final_status"] = existing.get("final_status") or merged.get("initial_match_status") or ""
    merged["final_value"] = existing.get("final_value") or merged.get("system_value") or ""
    return merged


def _resolve_report_image_url(file_key: str) -> str:
    request = Request(
        f"{CLAPPIA_FILE_DOWNLOAD_URL}?fileName={quote(file_key, safe='/')}",
        headers={"Accept": "text/plain"},
    )
    with urlopen(request, timeout=60) as response:
        resolved_url = response.read().decode().strip()
    if not resolved_url.startswith("http"):
        raise RuntimeError("Clappia file resolver did not return a valid signed URL.")
    return resolved_url


def refresh_report_image_urls(rows: list[dict[str, Any]], *, workers: int) -> list[dict[str, Any]]:
    if not rows:
        return rows

    def refresh_one(row: dict[str, Any]) -> dict[str, Any]:
        refreshed = dict(row)
        file_key = str(refreshed.get("file_key") or "").strip()
        if file_key:
            try:
                refreshed["report_image_url"] = _resolve_report_image_url(file_key)
            except Exception:
                refreshed["report_image_url"] = refreshed.get("report_image_url") or refreshed.get("image_url") or ""
        else:
            refreshed["report_image_url"] = refreshed.get("report_image_url") or refreshed.get("image_url") or ""
        return refreshed

    max_workers = max(1, min(workers, len(rows)))
    if max_workers == 1:
        return [refresh_one(row) for row in rows]

    refreshed_by_case_id: dict[str, dict[str, Any]] = {}
    with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
        future_map = {
            executor.submit(refresh_one, row): str(row["case_id"])
            for row in rows
        }
        for future in concurrent.futures.as_completed(future_map):
            refreshed_by_case_id[future_map[future]] = future.result()
    return [refreshed_by_case_id[str(row["case_id"])] for row in rows]


def _summary_status(row: dict[str, Any], *, adjusted_last_decimal: bool = False) -> str:
    status = str(row.get("final_status") or row.get("initial_match_status") or "").strip().lower()
    if status in {"correct", "mismatch", "needs_review", "blocked"}:
        resolved = status
    else:
        resolved = _match_status(
            str(row.get("system_status") or ""),
            str(row.get("matched") or "").lower() == "yes",
        )
    if adjusted_last_decimal and resolved == "mismatch" and _is_last_decimal_only_variation(
        str(row.get("expected_value") or ""),
        str(row.get("system_value") or ""),
        str(row.get("system_status") or ""),
    ):
        return "correct"
    return resolved


def _summary_counts(rows: list[dict[str, Any]], *, adjusted_last_decimal: bool = False) -> dict[str, Any]:
    total = len(rows)
    correct = sum(1 for row in rows if _summary_status(row, adjusted_last_decimal=adjusted_last_decimal) == "correct")
    mismatch = sum(1 for row in rows if _summary_status(row, adjusted_last_decimal=adjusted_last_decimal) == "mismatch")
    needs_review = sum(1 for row in rows if _summary_status(row, adjusted_last_decimal=adjusted_last_decimal) == "needs_review")
    blocked = sum(1 for row in rows if _summary_status(row, adjusted_last_decimal=adjusted_last_decimal) == "blocked")
    evaluated = total - blocked
    accuracy = round((correct / evaluated) * 100, 2) if evaluated else 0.0
    return {
        "total_cases": total,
        "correct": correct,
        "mismatch": mismatch,
        "needs_review": needs_review,
        "blocked": blocked,
        "evaluated": evaluated,
        "accuracy_percent": accuracy,
    }


def _guess_image_extension(url: str, content_type: str) -> str:
    guessed_from_type = mimetypes.guess_extension((content_type or "").split(";", 1)[0].strip())
    if guessed_from_type in {".jpe", ".jpeg", ".jpg", ".png", ".webp"}:
        return ".jpg" if guessed_from_type == ".jpe" else guessed_from_type
    path_suffix = Path(urlsplit(url).path).suffix.lower()
    if path_suffix in {".jpg", ".jpeg", ".png", ".webp"}:
        return path_suffix
    return ".jpg"


def cache_report_images(
    rows: list[dict[str, Any]],
    *,
    output_path: Path,
    workers: int,
) -> list[dict[str, Any]]:
    if not rows:
        return rows

    assets_dir = output_path.parent / f"{output_path.stem}_assets"
    assets_dir.mkdir(parents=True, exist_ok=True)

    def cache_one(index_and_row: tuple[int, dict[str, Any]]) -> dict[str, Any]:
        index, row = index_and_row
        cached = dict(row)
        image_url = str(cached.get("report_image_url") or cached.get("image_url") or "").strip()
        if not image_url.startswith("http"):
            return cached
        try:
            request = Request(image_url, headers={"Accept": "image/*"})
            with urlopen(request, timeout=120) as response:
                payload = response.read()
                extension = _guess_image_extension(image_url, response.headers.get("Content-Type", ""))
            asset_name = f"{index:04d}_{re.sub(r'[^A-Za-z0-9._-]+', '_', str(cached.get('case_id') or 'case'))}{extension}"
            asset_path = assets_dir / asset_name
            asset_path.write_bytes(payload)
            cached["report_image_local"] = f"{assets_dir.name}/{asset_name}"
        except Exception:
            cached["report_image_local"] = ""
        return cached

    indexed_rows = list(enumerate(rows, start=1))
    max_workers = max(1, min(workers, len(indexed_rows)))
    if max_workers == 1:
        return [cache_one(item) for item in indexed_rows]

    cached_by_case_id: dict[str, dict[str, Any]] = {}
    with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
        future_map = {
            executor.submit(cache_one, item): str(item[1]["case_id"])
            for item in indexed_rows
        }
        for future in concurrent.futures.as_completed(future_map):
            cached_by_case_id[future_map[future]] = future.result()
    return [cached_by_case_id[str(row["case_id"])] for row in rows]


def write_report_workbook(workbook_path: Path, case_rows: list[dict[str, Any]]) -> None:
    wb = Workbook()
    cases_ws = wb.active
    cases_ws.title = "cases"
    cases_ws.append(RESULT_HEADERS)
    for row in case_rows:
        cases_ws.append([row.get(header) for header in RESULT_HEADERS])

    summary_ws = wb.create_sheet("daily_summary")
    summary_headers = [
        "batch_day",
        "total_cases",
        "correct",
        "mismatch",
        "needs_review",
        "blocked",
        "accuracy_percent",
    ]
    summary_ws.append(summary_headers)

    grouped: dict[str, list[dict[str, Any]]] = {}
    for row in case_rows:
        grouped.setdefault(str(row["batch_day"]), []).append(row)
    for batch_day in sorted(grouped):
        rows = grouped[batch_day]
        counts = _summary_counts(rows)
        summary_ws.append([
            batch_day,
            counts["total_cases"],
            counts["correct"],
            counts["mismatch"],
            counts["needs_review"],
            counts["blocked"],
            counts["accuracy_percent"],
        ])

    overall_ws = wb.create_sheet("overall_summary")
    counts = _summary_counts(case_rows)
    overall_ws.append(["metric", "value"])
    overall_ws.append(["window_start", min((str(row["batch_day"]) for row in case_rows), default="")])
    overall_ws.append(["window_end", max((str(row["batch_day"]) for row in case_rows), default="")])
    overall_ws.append(["total_days", len(grouped)])
    overall_ws.append(["total_cases", counts["total_cases"]])
    overall_ws.append(["correct", counts["correct"]])
    overall_ws.append(["mismatch", counts["mismatch"]])
    overall_ws.append(["needs_review", counts["needs_review"]])
    overall_ws.append(["blocked", counts["blocked"]])
    overall_ws.append(["evaluated", counts["evaluated"]])
    overall_ws.append(["accuracy_percent", counts["accuracy_percent"]])
    overall_ws.append([])
    overall_ws.append(["section", "total_cases", "correct", "mismatch", "needs_review", "blocked", "accuracy_percent"])
    sections: dict[str, list[dict[str, Any]]] = {}
    for row in case_rows:
        sections.setdefault(str(row.get("section") or "Unknown Section"), []).append(row)
    for section in sorted(sections):
        section_counts = _summary_counts(sections[section])
        overall_ws.append([
            section,
            section_counts["total_cases"],
            section_counts["correct"],
            section_counts["mismatch"],
            section_counts["needs_review"],
            section_counts["blocked"],
            section_counts["accuracy_percent"],
        ])

    wb.save(workbook_path)


def _render_report_card(row: dict[str, Any], *, adjusted_last_decimal: bool = False) -> str:
        image_url = escape(str(row.get("report_image_local") or row.get("report_image_url") or row["image_url"]))
        original_image_url = escape(str(row["image_url"]))
        case_id = escape(str(row.get("case_id") or ""))
        batch_day = escape(str(row.get("batch_day") or ""))
        submission_id = escape(str(row["submission_id"]))
        case_label = escape(str(row.get("field") or row["case_label"]))
        section = escape(str(row["section"]))
        expected = escape(str(row["expected_display"] or row["expected_raw"] or ""))
        output = escape(str(row["system_value"] or row["system_status"]))
        reason = escape(str(row["reason"] or ""))
        initial_status = escape(str(row.get("initial_match_status") or ""))
        effective_status = _summary_status(row, adjusted_last_decimal=adjusted_last_decimal)
        final_status = escape(effective_status)
        review_status = escape(str(row.get("review_status") or ""))
        reviewed_value = escape(str(row.get("reviewed_value") or ""))
        review_note = escape(str(row.get("review_note") or ""))
        if adjusted_last_decimal and effective_status == "correct" and _is_last_decimal_only_variation(
            str(row.get("expected_value") or ""),
            str(row.get("system_value") or ""),
            str(row.get("system_status") or ""),
        ):
            result_label = "Correct (Last Decimal Accepted)"
        else:
            result_label = escape(str(row["result_label"] if effective_status != "correct" else "Correct"))
        created_at = escape(str(row["created_at"]))
        time_slot = escape(str(row["time_slot"] or ""))
        initial_status_value = escape(str(row.get("initial_match_status") or ""))
        final_status_value = escape(effective_status)
        return f"""
        <article class="card" data-case-id="{case_id}" data-day="{batch_day}" data-initial-status="{initial_status_value}" data-final-status="{final_status_value}">
          <div class="card-head">
            <span class="pill {'ok' if final_status == 'correct' else 'bad'} card-result-pill">{result_label}</span>
            <div class="meta">
              <h3>{case_label}</h3>
              <p>{section}</p>
              <p>Submission: <code>{submission_id}</code></p>
              <p>Created At: {created_at}</p>
              <p>Time Slot: {time_slot}</p>
            </div>
          </div>
          <a class="image-link" href="{image_url}" target="_blank" rel="noreferrer">Open image</a>
          <a class="image-link" href="{original_image_url}" target="_blank" rel="noreferrer">Open original Clappia link</a>
          <img class="photo" src="{image_url}" alt="{case_label}" loading="lazy" referrerpolicy="no-referrer" />
          <dl class="metrics">
            <div><dt>Expected</dt><dd>{expected}</dd></div>
            <div><dt>System Output</dt><dd>{output}</dd></div>
            <div><dt>System Status</dt><dd>{escape(str(row['system_status']))}</dd></div>
            <div><dt>Confidence</dt><dd>{escape(str(row['system_confidence']))}</dd></div>
            <div><dt>Initial Match Status</dt><dd>{initial_status}</dd></div>
            <div><dt>Final Status</dt><dd>{final_status}</dd></div>
            <div><dt>Review Status</dt><dd>{review_status}</dd></div>
            <div><dt>Reviewed Value</dt><dd>{reviewed_value}</dd></div>
            <div><dt>Review Note</dt><dd>{review_note}</dd></div>
            <div><dt>Reason</dt><dd>{reason}</dd></div>
          </dl>
          <section class="review-panel">
            <h4>Review Override</h4>
            <p class="review-help">Update only if this case should not count as a mismatch. The report accuracy updates immediately in the browser.</p>
            <div class="review-grid">
              <label>
                <span>Review Status</span>
                <select class="review-status">
                  <option value="">No override</option>
                  <option value="accepted_as_correct">Accepted as correct</option>
                  <option value="not_a_mismatch">Not a mismatch</option>
                  <option value="corrected_value">Corrected value</option>
                  <option value="ignored_case">Ignored case</option>
                  <option value="needs_review">Needs review</option>
                </select>
              </label>
              <label>
                <span>Reviewed Value</span>
                <input class="reviewed-value" type="text" value="{reviewed_value}" placeholder="Optional corrected value" />
              </label>
              <label class="review-note-wrap">
                <span>Review Note</span>
                <textarea class="review-note" rows="3" placeholder="Why this should not be counted as a mismatch">{review_note}</textarea>
              </label>
            </div>
            <div class="review-actions">
              <button class="reset-review" type="button">Reset Review</button>
              <span class="effective-status-chip"></span>
            </div>
          </section>
        </article>
        """


def _report_shell(*, title: str, lead: str, stats_html: str, body_html: str) -> str:
    return f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>{escape(title)}</title>
  <style>
    :root {{
      --bg: #f5f1e8;
      --panel: #fffdf8;
      --ink: #1f2a3a;
      --muted: #5d6b7a;
      --line: #d9d1c3;
      --ok: #d8f4df;
      --ok-ink: #1d6a35;
      --bad: #ffe1e1;
      --bad-ink: #8b2020;
      --shadow: 0 14px 40px rgba(32, 30, 24, 0.08);
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      background: linear-gradient(180deg, #f9f5ec 0%, var(--bg) 100%);
      color: var(--ink);
      font-family: Georgia, "Times New Roman", serif;
    }}
    main {{
      max-width: 1500px;
      margin: 0 auto;
      padding: 40px 32px 80px;
    }}
    h1 {{ font-size: 64px; line-height: 1; margin: 0 0 16px; }}
    p.lead {{ font-size: 18px; color: var(--muted); max-width: 980px; }}
    .stats {{
      display: grid;
      grid-template-columns: repeat(4, minmax(0, 1fr));
      gap: 20px;
      margin: 30px 0 50px;
    }}
    .stat {{
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 26px;
      padding: 28px 30px;
      box-shadow: var(--shadow);
    }}
    .stat h2 {{ margin: 0 0 10px; font-size: 18px; color: var(--muted); letter-spacing: 0.08em; text-transform: uppercase; }}
    .stat strong {{ font-size: 54px; }}
    details {{
      background: transparent;
      margin: 24px 0;
      border-top: 2px solid var(--line);
      padding-top: 18px;
    }}
    summary {{
      cursor: pointer;
      font-size: 42px;
      font-weight: 700;
      list-style: none;
    }}
    summary::-webkit-details-marker {{ display: none; }}
    .cards {{
      display: grid;
      grid-template-columns: repeat(2, minmax(0, 1fr));
      gap: 28px;
      margin-top: 24px;
    }}
    .card {{
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 28px;
      padding: 28px;
      box-shadow: var(--shadow);
    }}
    .card-head {{
      display: flex;
      gap: 18px;
      align-items: flex-start;
      margin-bottom: 20px;
    }}
    .meta h3 {{ margin: 0 0 8px; font-size: 30px; }}
    .meta p {{ margin: 6px 0; color: var(--muted); font-size: 17px; }}
    .pill {{
      display: inline-block;
      border-radius: 999px;
      padding: 10px 16px;
      font-size: 14px;
      font-weight: 700;
      letter-spacing: 0.08em;
      text-transform: uppercase;
      white-space: nowrap;
    }}
    .pill.ok {{ background: var(--ok); color: var(--ok-ink); }}
    .pill.bad {{ background: var(--bad); color: var(--bad-ink); }}
    .image-link {{
      display: inline-block;
      margin-bottom: 14px;
      color: #0b4f8a;
      text-decoration: none;
      font-weight: 700;
      margin-right: 16px;
    }}
    .photo {{
      width: 100%;
      height: 520px;
      object-fit: contain;
      background: #ece8df;
      border-radius: 18px;
      border: 1px solid var(--line);
      display: block;
      margin-bottom: 20px;
    }}
    .metrics {{
      display: grid;
      grid-template-columns: 1fr 1fr;
      gap: 14px 28px;
      margin: 0;
    }}
    .metrics div {{
      display: grid;
      grid-template-columns: 180px 1fr;
      gap: 12px;
    }}
    .metrics dt {{ font-weight: 700; }}
    .metrics dd {{ margin: 0; color: var(--muted); word-break: break-word; }}
    .subsection-title {{
      margin: 20px 0 0;
      font-size: 26px;
    }}
    .toolbar {{
      display: flex;
      flex-wrap: wrap;
      gap: 12px;
      margin: 8px 0 24px;
      align-items: center;
    }}
    .toolbar button, .toolbar label.button-like {{
      border: 1px solid var(--line);
      background: var(--panel);
      color: var(--ink);
      padding: 10px 16px;
      border-radius: 999px;
      cursor: pointer;
      font: inherit;
      box-shadow: var(--shadow);
    }}
    .toolbar input[type="file"] {{ display: none; }}
    .toolbar-note {{
      color: var(--muted);
      font-size: 15px;
    }}
    .review-panel {{
      margin-top: 24px;
      padding-top: 20px;
      border-top: 1px solid var(--line);
    }}
    .review-panel h4 {{
      margin: 0 0 8px;
      font-size: 22px;
    }}
    .review-help {{
      margin: 0 0 14px;
      color: var(--muted);
      font-size: 15px;
    }}
    .review-grid {{
      display: grid;
      grid-template-columns: 1fr 1fr;
      gap: 14px 18px;
    }}
    .review-grid label {{
      display: grid;
      gap: 8px;
      font-size: 14px;
      font-weight: 700;
      color: var(--ink);
    }}
    .review-note-wrap {{
      grid-column: 1 / -1;
    }}
    .review-grid select,
    .review-grid input,
    .review-grid textarea {{
      width: 100%;
      border: 1px solid var(--line);
      background: #fff;
      color: var(--ink);
      border-radius: 12px;
      padding: 10px 12px;
      font: inherit;
    }}
    .review-actions {{
      display: flex;
      gap: 12px;
      align-items: center;
      margin-top: 14px;
    }}
    .reset-review {{
      border: 1px solid var(--line);
      background: transparent;
      color: var(--ink);
      padding: 8px 12px;
      border-radius: 999px;
      cursor: pointer;
      font: inherit;
    }}
    .effective-status-chip {{
      display: inline-block;
      border-radius: 999px;
      padding: 8px 12px;
      font-size: 13px;
      font-weight: 700;
      letter-spacing: 0.04em;
      text-transform: uppercase;
      background: #ece8df;
      color: var(--ink);
    }}
    .effective-status-chip.ok {{
      background: var(--ok);
      color: var(--ok-ink);
    }}
    .effective-status-chip.bad {{
      background: var(--bad);
      color: var(--bad-ink);
    }}
    .effective-status-chip.neutral {{
      background: #ece8df;
      color: var(--ink);
    }}
    .day-summary-line {{
      display: flex;
      flex-wrap: wrap;
      gap: 14px;
      align-items: center;
    }}
    .day-summary-line .meta {{
      font-size: 18px;
      color: var(--muted);
      font-weight: 400;
    }}
    @media (max-width: 1100px) {{
      .stats, .cards {{ grid-template-columns: 1fr; }}
      h1 {{ font-size: 44px; }}
      summary {{ font-size: 32px; }}
      .photo {{ height: 380px; }}
      .metrics div {{ grid-template-columns: 1fr; }}
      .review-grid {{ grid-template-columns: 1fr; }}
    }}
  </style>
</head>
<body>
  <main>
    <h1>{escape(title)}</h1>
    <p class="lead">{lead}</p>
    <section class="toolbar">
      <button id="export-reviews" type="button">Export Reviews</button>
      <label class="button-like" for="import-reviews">Import Reviews</label>
      <input id="import-reviews" type="file" accept="application/json" />
      <button id="clear-reviews" type="button">Clear All Reviews</button>
      <span class="toolbar-note">Reviews are stored in this browser and can be exported/imported as JSON.</span>
    </section>
    {stats_html}
    {body_html}
  </main>
  <script>
    (() => {{
      const storageKey = "mfcvision-history-review-" + {json.dumps(title)};
      const cards = Array.from(document.querySelectorAll(".card[data-case-id]"));
      const summaryCards = {{
        total: document.querySelector('[data-summary="total"] strong'),
        correct: document.querySelector('[data-summary="correct"] strong'),
        mismatch: document.querySelector('[data-summary="mismatch"] strong'),
        needsReview: document.querySelector('[data-summary="needs_review"] strong'),
        blocked: document.querySelector('[data-summary="blocked"] strong'),
        evaluated: document.querySelector('[data-summary="evaluated"] strong'),
        accuracy: document.querySelector('[data-summary="accuracy"] strong'),
      }};
      const dayStats = new Map();
      document.querySelectorAll("[data-day-stats]").forEach((el) => {{
        dayStats.set(el.getAttribute("data-day-stats"), el);
      }});

      function loadReviews() {{
        try {{
          return JSON.parse(localStorage.getItem(storageKey) || "{{}}");
        }} catch (error) {{
          return {{}};
        }}
      }}

      let reviews = loadReviews();

      function saveReviews() {{
        localStorage.setItem(storageKey, JSON.stringify(reviews));
      }}

      function normalizeReview(review) {{
        return {{
          review_status: String(review?.review_status || ""),
          reviewed_value: String(review?.reviewed_value || ""),
          review_note: String(review?.review_note || ""),
        }};
      }}

      function effectiveStatus(initialStatus, reviewStatus) {{
        if (reviewStatus === "accepted_as_correct" || reviewStatus === "not_a_mismatch" || reviewStatus === "corrected_value") {{
          return "correct";
        }}
        if (reviewStatus === "needs_review") {{
          return "needs_review";
        }}
        if (reviewStatus === "ignored_case") {{
          return "ignored";
        }}
        return initialStatus || "mismatch";
      }}

      function resultLabelFor(status) {{
        if (status === "correct") return "Reviewed Correct";
        if (status === "mismatch") return "Mismatch";
        if (status === "needs_review") return "Needs Review";
        if (status === "blocked") return "Blocked";
        if (status === "ignored") return "Ignored";
        return status || "Unknown";
      }}

      function chipClassFor(status) {{
        if (status === "correct") return "ok";
        if (status === "mismatch" || status === "blocked") return "bad";
        return "neutral";
      }}

      function applyReviewToCard(card) {{
        const caseId = card.dataset.caseId;
        const initialStatus = card.dataset.finalStatus || card.dataset.initialStatus || "";
        const review = normalizeReview(reviews[caseId] || {{}});
        const select = card.querySelector(".review-status");
        const reviewedValue = card.querySelector(".reviewed-value");
        const reviewNote = card.querySelector(".review-note");
        const pill = card.querySelector(".card-result-pill");
        const chip = card.querySelector(".effective-status-chip");
        if (select && select.value !== review.review_status) select.value = review.review_status;
        if (reviewedValue && reviewedValue.value !== review.reviewed_value) reviewedValue.value = review.reviewed_value;
        if (reviewNote && reviewNote.value !== review.review_note) reviewNote.value = review.review_note;
        const status = effectiveStatus(initialStatus, review.review_status);
        card.dataset.effectiveStatus = status;
        if (pill) {{
          pill.textContent = resultLabelFor(status);
          pill.classList.remove("ok", "bad");
          pill.classList.add(status === "correct" ? "ok" : "bad");
        }}
        if (chip) {{
          chip.textContent = "Effective: " + resultLabelFor(status);
          chip.className = "effective-status-chip " + chipClassFor(status);
        }}
      }}

      function recalcSummaries() {{
        const totals = {{ total: 0, correct: 0, mismatch: 0, needs_review: 0, blocked: 0, ignored: 0 }};
        const byDay = new Map();
        cards.forEach((card) => {{
          const day = card.dataset.day || "";
          const status = card.dataset.effectiveStatus || card.dataset.finalStatus || card.dataset.initialStatus || "mismatch";
          totals.total += 1;
          if (status in totals) totals[status] += 1;
          if (!byDay.has(day)) byDay.set(day, {{ total: 0, correct: 0, mismatch: 0, needs_review: 0, blocked: 0, ignored: 0 }});
          const bucket = byDay.get(day);
          bucket.total += 1;
          if (status in bucket) bucket[status] += 1;
        }});

        const evaluated = totals.total - totals.blocked - totals.ignored;
        const accuracy = evaluated ? ((totals.correct / evaluated) * 100).toFixed(2) : "0.00";
        if (summaryCards.total) summaryCards.total.textContent = String(totals.total);
        if (summaryCards.correct) summaryCards.correct.textContent = String(totals.correct);
        if (summaryCards.mismatch) summaryCards.mismatch.textContent = String(totals.mismatch);
        if (summaryCards.needsReview) summaryCards.needsReview.textContent = String(totals.needs_review);
        if (summaryCards.blocked) summaryCards.blocked.textContent = String(totals.blocked);
        if (summaryCards.evaluated) summaryCards.evaluated.textContent = String(evaluated);
        if (summaryCards.accuracy) summaryCards.accuracy.textContent = accuracy + "%";

        byDay.forEach((counts, day) => {{
          const evaluatedDay = counts.total - counts.blocked - counts.ignored;
          const accuracyDay = evaluatedDay ? ((counts.correct / evaluatedDay) * 100).toFixed(2) : "0.00";
          const container = dayStats.get(day);
          if (!container) return;
          container.querySelector('[data-day-metric="total"]').textContent = String(counts.total);
          container.querySelector('[data-day-metric="correct"]').textContent = String(counts.correct);
          container.querySelector('[data-day-metric="mismatch"]').textContent = String(counts.mismatch);
          container.querySelector('[data-day-metric="needs_review"]').textContent = String(counts.needs_review);
          container.querySelector('[data-day-metric="blocked"]').textContent = String(counts.blocked);
          container.querySelector('[data-day-metric="accuracy"]').textContent = accuracyDay + "%";
          const summaryLine = document.querySelector('[data-day-summary="' + CSS.escape(day) + '"]');
          if (summaryLine) {{
            summaryLine.querySelector('[data-day-summary-metric="accuracy"]').textContent = accuracyDay + "%";
            summaryLine.querySelector('[data-day-summary-metric="correct"]').textContent = String(counts.correct);
            summaryLine.querySelector('[data-day-summary-metric="evaluated"]').textContent = String(evaluatedDay);
          }}
        }});
      }}

      function persistCard(card) {{
        const caseId = card.dataset.caseId;
        const select = card.querySelector(".review-status");
        const reviewedValue = card.querySelector(".reviewed-value");
        const reviewNote = card.querySelector(".review-note");
        reviews[caseId] = normalizeReview({{
          review_status: select?.value || "",
          reviewed_value: reviewedValue?.value || "",
          review_note: reviewNote?.value || "",
        }});
        if (!reviews[caseId].review_status && !reviews[caseId].reviewed_value && !reviews[caseId].review_note) {{
          delete reviews[caseId];
        }}
        saveReviews();
        applyReviewToCard(card);
        recalcSummaries();
      }}

      cards.forEach((card) => {{
        applyReviewToCard(card);
        card.querySelector(".review-status")?.addEventListener("change", () => persistCard(card));
        card.querySelector(".reviewed-value")?.addEventListener("input", () => persistCard(card));
        card.querySelector(".review-note")?.addEventListener("input", () => persistCard(card));
        card.querySelector(".reset-review")?.addEventListener("click", () => {{
          delete reviews[card.dataset.caseId];
          saveReviews();
          applyReviewToCard(card);
          recalcSummaries();
        }});
      }});

      document.getElementById("clear-reviews")?.addEventListener("click", () => {{
        reviews = {{}};
        saveReviews();
        cards.forEach(applyReviewToCard);
        recalcSummaries();
      }});

      document.getElementById("export-reviews")?.addEventListener("click", () => {{
        const payload = {{
          exported_at: new Date().toISOString(),
          report_title: {json.dumps(title)},
          reviews,
        }};
        const blob = new Blob([JSON.stringify(payload, null, 2)], {{ type: "application/json" }});
        const url = URL.createObjectURL(blob);
        const link = document.createElement("a");
        link.href = url;
        link.download = "mfcvision-report-reviews.json";
        document.body.appendChild(link);
        link.click();
        link.remove();
        URL.revokeObjectURL(url);
      }});

      document.getElementById("import-reviews")?.addEventListener("change", async (event) => {{
        const file = event.target.files?.[0];
        if (!file) return;
        try {{
          const text = await file.text();
          const payload = JSON.parse(text);
          reviews = payload.reviews || {{}};
          saveReviews();
          cards.forEach(applyReviewToCard);
          recalcSummaries();
        }} catch (error) {{
          window.alert("Could not import review JSON.");
        }}
        event.target.value = "";
      }});

      recalcSummaries();
    }})();
  </script>
</body>
</html>
"""


def build_day_report_html(
    *,
    batch_day: str,
    rows: list[dict[str, Any]],
    output_path: Path,
    adjusted_last_decimal: bool = False,
) -> None:
    counts = _summary_counts(rows, adjusted_last_decimal=adjusted_last_decimal)
    worked = [row for row in rows if _summary_status(row, adjusted_last_decimal=adjusted_last_decimal) == "correct"]
    failed = [row for row in rows if _summary_status(row, adjusted_last_decimal=adjusted_last_decimal) != "correct"]

    stats_html = f"""
    <section class="stats">
      <article class="stat" data-summary="total"><h2>Total Cases</h2><strong>{counts['total_cases']}</strong></article>
      <article class="stat" data-summary="correct"><h2>Correct</h2><strong>{counts['correct']}</strong></article>
      <article class="stat" data-summary="mismatch"><h2>Did Not Match</h2><strong>{counts['mismatch']}</strong></article>
      <article class="stat" data-summary="accuracy"><h2>Accuracy</h2><strong>{counts['accuracy_percent']}%</strong></article>
    </section>
    <section class="stats">
      <article class="stat" data-summary="needs_review"><h2>Needs Review</h2><strong>{counts['needs_review']}</strong></article>
      <article class="stat" data-summary="blocked"><h2>Blocked</h2><strong>{counts['blocked']}</strong></article>
      <article class="stat" data-summary="evaluated"><h2>Evaluated</h2><strong>{counts['evaluated']}</strong></article>
      <article class="stat"><h2>Batch Day</h2><strong style="font-size:34px">{escape(batch_day)}</strong></article>
    </section>
    """
    body_html = f"""
    <details open>
      <summary>Worked ({len(worked)})</summary>
      <div class="cards">
        {''.join(_render_report_card(row, adjusted_last_decimal=adjusted_last_decimal) for row in worked)}
      </div>
    </details>
    <details open>
      <summary>Did Not Work / Needs Review ({len(failed)})</summary>
      <div class="cards">
        {''.join(_render_report_card(row, adjusted_last_decimal=adjusted_last_decimal) for row in failed)}
      </div>
    </details>
    """
    output_path.write_text(
        _report_shell(
            title=f"History Batch Report {batch_day}",
            lead=(
                f"Batch day <code>{escape(batch_day)}</code>. This report keeps the original Clappia wrapper link on each card for traceability. "
                + ("Last-decimal-only mismatches are counted as correct. " if adjusted_last_decimal else "")
                + ("Images are cached locally with this report." if any(row.get('report_image_local') for row in rows) else "This report uses fresh signed Clappia file URLs for image display.")
            ),
            stats_html=stats_html,
            body_html=body_html,
        )
    )


def build_window_report_html(
    *,
    window_start: str,
    window_end: str,
    rows: list[dict[str, Any]],
    output_path: Path,
) -> None:
    counts = _summary_counts(rows)
    grouped: dict[str, list[dict[str, Any]]] = {}
    for row in rows:
        grouped.setdefault(str(row["batch_day"]), []).append(row)

    sections_html: list[str] = []
    for batch_day in sorted(grouped):
        day_rows = grouped[batch_day]
        day_counts = _summary_counts(day_rows)
        worked = [row for row in day_rows if _summary_status(row) == "correct"]
        failed = [row for row in day_rows if _summary_status(row) != "correct"]
        sections_html.append(
            f"""
            <details>
              <summary>
                <span class="day-summary-line" data-day-summary="{escape(batch_day)}">
                  <span>{escape(batch_day)}</span>
                  <span class="meta">Accuracy <span data-day-summary-metric="accuracy">{day_counts['accuracy_percent']}%</span></span>
                  <span class="meta">Correct <span data-day-summary-metric="correct">{day_counts['correct']}</span> / <span data-day-summary-metric="evaluated">{day_counts['evaluated']}</span></span>
                </span>
              </summary>
              <section class="stats" data-day-stats="{escape(batch_day)}">
                <article class="stat"><h2>Total Cases</h2><strong data-day-metric="total">{day_counts['total_cases']}</strong></article>
                <article class="stat"><h2>Correct</h2><strong data-day-metric="correct">{day_counts['correct']}</strong></article>
                <article class="stat"><h2>Mismatch</h2><strong data-day-metric="mismatch">{day_counts['mismatch']}</strong></article>
                <article class="stat"><h2>Needs Review</h2><strong data-day-metric="needs_review">{day_counts['needs_review']}</strong></article>
                <article class="stat"><h2>Blocked</h2><strong data-day-metric="blocked">{day_counts['blocked']}</strong></article>
                <article class="stat"><h2>Accuracy</h2><strong data-day-metric="accuracy">{day_counts['accuracy_percent']}%</strong></article>
              </section>
              <h3 class="subsection-title">Worked ({len(worked)})</h3>
              <div class="cards">
                {''.join(_render_report_card(row) for row in worked)}
              </div>
              <h3 class="subsection-title">Did Not Work / Needs Review ({len(failed)})</h3>
              <div class="cards">
                {''.join(_render_report_card(row) for row in failed)}
              </div>
            </details>
            """
        )

    stats_html = f"""
    <section class="stats">
      <article class="stat" data-summary="total"><h2>Total Cases</h2><strong>{counts['total_cases']}</strong></article>
      <article class="stat" data-summary="correct"><h2>Correct</h2><strong>{counts['correct']}</strong></article>
      <article class="stat" data-summary="mismatch"><h2>Did Not Match</h2><strong>{counts['mismatch']}</strong></article>
      <article class="stat" data-summary="accuracy"><h2>Accuracy</h2><strong>{counts['accuracy_percent']}%</strong></article>
    </section>
    <section class="stats">
      <article class="stat" data-summary="needs_review"><h2>Needs Review</h2><strong>{counts['needs_review']}</strong></article>
      <article class="stat" data-summary="blocked"><h2>Blocked</h2><strong>{counts['blocked']}</strong></article>
      <article class="stat" data-summary="evaluated"><h2>Evaluated</h2><strong>{counts['evaluated']}</strong></article>
      <article class="stat"><h2>Days</h2><strong style="font-size:34px">{len(grouped)}</strong></article>
    </section>
    """
    output_path.write_text(
        _report_shell(
            title="Last 2 Weeks Validation Report",
            lead=(
                f"Window <code>{escape(window_start)}</code> to <code>{escape(window_end)}</code>. "
                "Each day is collapsible. The report uses fresh signed Clappia file URLs for image display and keeps the original Clappia wrapper link on each card."
            ),
            stats_html=stats_html,
            body_html="".join(sections_html),
        )
    )


def write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2, ensure_ascii=False))


def _sort_case_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return sorted(
        rows,
        key=lambda row: (
            str(row.get("batch_day") or ""),
            str(row.get("created_at") or ""),
            str(row.get("submission_id") or ""),
            str(row.get("case_label") or ""),
        ),
    )


def _build_window_summary_payload(
    *,
    window_start: str,
    window_end: str,
    rows: list[dict[str, Any]],
    target_days: list[str],
    processed_days: list[str],
    report_xlsx: Path,
    overall_json_path: Path,
    overall_html_path: Path,
    results_dir: Path,
    workers: int,
) -> dict[str, Any]:
    return {
        "window_start": window_start,
        "window_end": window_end,
        "days": target_days,
        "processed_days": processed_days,
        "pending_days": [day for day in target_days if day not in set(processed_days)],
        "counts": _summary_counts(rows),
        "report_xlsx": str(report_xlsx),
        "overall_json": str(overall_json_path),
        "overall_html": str(overall_html_path),
        "results_dir": str(results_dir),
        "workers": workers,
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="Run history validation from the exported Clappia workbook.")
    parser.add_argument("--source-xlsx", required=True, type=Path)
    parser.add_argument("--report-xlsx", required=True, type=Path)
    parser.add_argument("--results-json", type=Path, default=None)
    parser.add_argument("--report-html", type=Path, default=None)
    parser.add_argument("--manifest-json", type=Path, default=None)
    parser.add_argument("--day", default=None, help="ISO day to run, for example 2026-03-10. Defaults to Day 1 of the last 2-week window.")
    parser.add_argument("--all-days", action="store_true", help="Run the full last 2-week window day by day.")
    parser.add_argument("--results-dir", type=Path, default=None, help="Directory for per-day manifests/results/reports when running all days.")
    parser.add_argument("--overall-json", type=Path, default=None, help="Combined case results JSON for the full last 2-week run.")
    parser.add_argument("--overall-html", type=Path, default=None, help="Combined HTML report for the full last 2-week run.")
    parser.add_argument("--summary-json", type=Path, default=None, help="Window summary JSON output.")
    parser.add_argument("--workers", type=int, default=8, help="Number of parallel worker threads for case execution.")
    parser.add_argument("--rebuild-results-json", type=Path, default=None, help="Rebuild report HTML from an existing day/window results JSON without rerunning Gemini.")
    parser.add_argument("--cache-report-images", action="store_true", help="Download report images locally so the HTML remains viewable after signed URLs expire.")
    parser.add_argument("--adjust-last-decimal", action="store_true", help="Treat last-decimal-only mismatches as correct in the rendered report.")
    args = parser.parse_args()

    if args.rebuild_results_json:
        results_json_path = args.rebuild_results_json.resolve()
        rows = json.loads(results_json_path.read_text())
        if not isinstance(rows, list):
            raise SystemExit("Expected the rebuild results JSON to contain a list of case rows.")
        normalized_rows = [_normalize_existing_case_row(row) for row in rows]
        refreshed_rows = refresh_report_image_urls(normalized_rows, workers=args.workers)
        report_html_path = (args.report_html or results_json_path.with_name(results_json_path.name.replace("_results.json", "_report.html"))).resolve()
        if args.cache_report_images:
            refreshed_rows = cache_report_images(refreshed_rows, output_path=report_html_path, workers=args.workers)
        write_json(results_json_path, refreshed_rows)
        batch_days = sorted({str(row.get("batch_day") or "") for row in refreshed_rows if str(row.get("batch_day") or "").strip()})
        if len(batch_days) == 1:
            build_day_report_html(
                batch_day=batch_days[0],
                rows=refreshed_rows,
                output_path=report_html_path,
                adjusted_last_decimal=args.adjust_last_decimal,
            )
            print(
                json.dumps(
                    {
                        "batch_day": batch_days[0],
                        "counts": _summary_counts(refreshed_rows, adjusted_last_decimal=args.adjust_last_decimal),
                        "results_json": str(results_json_path),
                        "report_html": str(report_html_path),
                        "cache_report_images": args.cache_report_images,
                        "adjust_last_decimal": args.adjust_last_decimal,
                    },
                    indent=2,
                )
            )
            return 0
        raise SystemExit("Rebuild mode currently supports a single-day results JSON.")

    source_xlsx = args.source_xlsx.resolve()
    all_cases = parse_cases_from_workbook(source_xlsx)
    if not all_cases:
        raise SystemExit("No testable cases found in workbook.")

    window_days = select_last_two_week_days(all_cases)
    if not window_days:
        raise SystemExit("Could not derive the last 2-week window from Created At.")

    if args.all_days and args.day:
        raise SystemExit("Use either --day for a single batch or --all-days for the full window, not both.")

    if args.all_days:
        target_days = window_days
        results_dir = (args.results_dir or (args.report_xlsx.parent / "last_2_weeks_outputs")).resolve()
        results_dir.mkdir(parents=True, exist_ok=True)
        overall_json_path = (args.overall_json or (results_dir / "last_2_weeks_case_results.json")).resolve()
        overall_html_path = (args.overall_html or (results_dir / "last_2_weeks_report.html")).resolve()
        summary_json_path = (args.summary_json or (results_dir / "last_2_weeks_summary.json")).resolve()
    else:
        target_day = args.day or window_days[0]
        if target_day not in window_days:
            raise SystemExit(
                f"Requested day {target_day} is not in the last 2-week window: {window_days[0]} to {window_days[-1]}"
            )
        target_days = [target_day]
        results_dir = None
        overall_json_path = None
        overall_html_path = None
        summary_json_path = None

    existing_rows = [_normalize_existing_case_row(row) for row in load_existing_case_rows(args.report_xlsx)]
    target_days_set = set(target_days)
    existing_by_case_id = {str(row.get("case_id")): row for row in existing_rows}
    incremental_rows_by_case_id = {
        str(row["case_id"]): row
        for row in existing_rows
        if str(row.get("batch_day")) not in target_days_set
    }

    per_day_results: dict[str, list[dict[str, Any]]] = {}
    processed_days: list[str] = []
    for batch_day in target_days:
        day_cases = [case for case in all_cases if case.batch_day == batch_day]
        if not day_cases:
            continue

        if args.all_days:
            manifest_path = results_dir / f"day_{batch_day}_manifest.json"
            results_path = results_dir / f"day_{batch_day}_results.json"
            report_path = results_dir / f"day_{batch_day}_report.html"
        else:
            manifest_path = (args.manifest_json or (args.report_xlsx.parent / f"day_{batch_day}_manifest.json")).resolve()
            results_path = (args.results_json or (args.report_xlsx.parent / f"day_{batch_day}_results.json")).resolve()
            report_path = (args.report_html or (args.report_xlsx.parent / f"day_{batch_day}_report.html")).resolve()

        write_json(
            manifest_path,
            {
                "source_workbook": str(source_xlsx),
                "window_start": window_days[0],
                "window_end": window_days[-1],
                "batch_day": batch_day,
                "case_count": len(day_cases),
                "cases": [asdict(case) | {"case_id": case.case_id} for case in day_cases],
            },
        )

        day_results = run_batch(
            day_cases,
            source_workbook=source_xlsx.name,
            window_start=window_days[0],
            window_end=window_days[-1],
            workers=args.workers,
        )
        per_day_results[batch_day] = day_results
        processed_days.append(batch_day)

        refreshed_day_rows = refresh_report_image_urls(day_results, workers=args.workers)
        merged_day_rows = [
            _merge_review_fields(row, existing_by_case_id.get(str(row["case_id"])))
            for row in refreshed_day_rows
        ]
        for row in merged_day_rows:
            incremental_rows_by_case_id[str(row["case_id"])] = row

        incremental_combined_rows = _sort_case_rows(list(incremental_rows_by_case_id.values()))
        write_report_workbook(args.report_xlsx, incremental_combined_rows)
        write_json(results_path, refreshed_day_rows)

        if args.all_days:
            current_window_rows = [
                row for row in incremental_combined_rows if str(row.get("batch_day")) in set(processed_days)
            ]
            write_json(overall_json_path, current_window_rows)
            write_json(
                summary_json_path,
                _build_window_summary_payload(
                    window_start=window_days[0],
                    window_end=window_days[-1],
                    rows=current_window_rows,
                    target_days=target_days,
                    processed_days=processed_days,
                    report_xlsx=args.report_xlsx,
                    overall_json_path=overall_json_path,
                    overall_html_path=overall_html_path,
                    results_dir=results_dir,
                    workers=args.workers,
                ),
            )

    refreshed_day_results: dict[str, list[dict[str, Any]]] = {}
    for batch_day in target_days:
        day_rows = per_day_results.get(batch_day, [])
        if not day_rows:
            continue
        refreshed_day_results[batch_day] = refresh_report_image_urls(day_rows, workers=args.workers)

    merged_new_rows: list[dict[str, Any]] = []
    for batch_day in target_days:
        for row in refreshed_day_results.get(batch_day, []):
            merged_new_rows.append(_merge_review_fields(row, existing_by_case_id.get(str(row["case_id"]))))

    kept_rows = [row for row in existing_rows if str(row.get("batch_day")) not in target_days_set]
    combined_rows = kept_rows + merged_new_rows
    combined_rows = _sort_case_rows(combined_rows)
    write_report_workbook(args.report_xlsx, combined_rows)

    for batch_day in target_days:
        day_rows = refreshed_day_results.get(batch_day, [])
        if args.all_days:
            results_path = results_dir / f"day_{batch_day}_results.json"
            report_path = results_dir / f"day_{batch_day}_report.html"
        else:
            results_path = (args.results_json or (args.report_xlsx.parent / f"day_{batch_day}_results.json")).resolve()
            report_path = (args.report_html or (args.report_xlsx.parent / f"day_{batch_day}_report.html")).resolve()
        write_json(results_path, day_rows)
        build_day_report_html(batch_day=batch_day, rows=day_rows, output_path=report_path)

    if args.all_days:
        window_rows = [row for row in combined_rows if str(row.get("batch_day")) in set(target_days)]
        write_json(overall_json_path, window_rows)
        build_window_report_html(
            window_start=window_days[0],
            window_end=window_days[-1],
            rows=window_rows,
            output_path=overall_html_path,
        )
        summary_payload = _build_window_summary_payload(
            window_start=window_days[0],
            window_end=window_days[-1],
            rows=window_rows,
            target_days=target_days,
            processed_days=target_days,
            report_xlsx=args.report_xlsx,
            overall_json_path=overall_json_path,
            overall_html_path=overall_html_path,
            results_dir=results_dir,
            workers=args.workers,
        )
        write_json(summary_json_path, summary_payload)
        print(json.dumps(summary_payload, indent=2))
    else:
        batch_day = target_days[0]
        day_rows = refreshed_day_results.get(batch_day, [])
        counts = _summary_counts(day_rows)
        print(
            json.dumps(
                {
                    "batch_day": batch_day,
                    "window_start": window_days[0],
                    "window_end": window_days[-1],
                    "case_count": counts["total_cases"],
                    "correct": counts["correct"],
                    "mismatch": counts["mismatch"],
                    "needs_review": counts["needs_review"],
                    "blocked": counts["blocked"],
                    "evaluated": counts["evaluated"],
                    "accuracy_percent": counts["accuracy_percent"],
                    "report_xlsx": str(args.report_xlsx),
                    "results_json": str((args.results_json or (args.report_xlsx.parent / f"day_{batch_day}_results.json")).resolve()),
                    "report_html": str((args.report_html or (args.report_xlsx.parent / f"day_{batch_day}_report.html")).resolve()),
                    "manifest_json": str((args.manifest_json or (args.report_xlsx.parent / f"day_{batch_day}_manifest.json")).resolve()),
                    "workers": args.workers,
                },
                indent=2,
            )
        )
    return 0


if __name__ == "__main__":
    sys.exit(main())
